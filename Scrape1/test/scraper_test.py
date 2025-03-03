import unittest
from unittest.mock import patch
import requests
from requests.models import Response
import xlsxwriter
import pandas as pd
from openpyxl import load_workbook
import os
import sys

# Get the absolute path of the src folder and add it to sys.path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../')))

from src.scraper import Scraper
from src.exceptions.failed_response_exception import FailedResponseException

URL_TO_SCRAPE = "https://books.toscrape.com/"
FILENAME = os.path.abspath(os.path.join(os.path.dirname(__file__), "../books_data.xlsx"))

class TestScraperMethods(unittest.TestCase):
    @classmethod
    def setUpClass(self):
        """Runs once before any tests execute."""
        self.scraper = Scraper(URL_TO_SCRAPE)

        # Ensure Excel file is created before tests
        books_data = [("book1", "genre1", "$1.00"),
                      ("book2", "genre2", "$2.00"),
                      ("book3", "genre3", "$3.00"),
                      ("book4", "genre4", "$4.00"),
                      ("book5", "genre5", "$5.00"),
                      ("book6", "genre6", "$6.00"),
                      ("book7", "genre7", "$7.00"),
                      ("book8", "genre8", "$8.00"),
                      ("book9", "genre9", "$9.00")]
        
        self.scraper.write_data_to_excel(books_data)

    @patch("src.scraper.requests.get")
    def test_connect_success(self, mock_get):
        mock_response = Response()
        mock_response.status_code = 200
        mock_response._content = b"Fake response content"
        mock_get.return_value = mock_response

        response = self.scraper.connect(self.scraper.base_url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.text, "Fake response content")

    @patch("src.scraper.requests.get")
    def test_connect_failure(self, mock_get):
        mock_response = Response()
        mock_response.status_code = 404
        mock_get.return_value = mock_response

        with self.assertRaises(FailedResponseException) as context:
            self.scraper.connect(self.scraper.base_url)

        self.assertIn(
            "Could not connect to server... Server returned with error code 404",
            str(context.exception).replace("\n", "").strip()
        )

    def test_genres_extraction(self):
        mock_response = Response()
        mock_response.status_code = 200
        mock_response._content = b"""
            <div class="side_categories">
                <ul>
                    <li>
                        <ul>
                            <li><a href="/category/books/travel_2/index.html">Travel</a></li>
                            <li><a href="/category/books/crime_51/index.html">Crime</a></li>
                        </ul>
                    </li>
                </ul>
            </div>
        """

        categories, category_links = self.scraper.get_genres(mock_response)
        
        self.assertEqual(categories, ["Travel", "Crime"])
        self.assertEqual(category_links["Travel"], "https://books.toscrape.com/category/books/travel_2/index.html")
        self.assertEqual(category_links["Crime"], "https://books.toscrape.com/category/books/crime_51/index.html")

    def test_excel_creation(self):
        """Check if the correct file exist."""
        self.assertTrue(os.path.exists(FILENAME))

    def test_excel_sheet_names(self):
        """Check if the correct sheets exist."""
        workbook = load_workbook(FILENAME)
        self.assertIn("Books_Data", workbook.sheetnames)

    def test_excel_cell_values(self):
        """Verify specific cell values in the Excel file."""
        workbook = load_workbook(FILENAME)
        sheet = workbook["Books_Data"]

        # Check header row
        self.assertEqual(sheet.cell(row=1, column=1).value, "Title")
        self.assertEqual(sheet.cell(row=1, column=2).value, "Genre")
        self.assertEqual(sheet.cell(row=1, column=3).value, "Price")

        # Check data values
        for i in range(1, 9):
            self.assertEqual(sheet.cell(row=(i+1), column=1).value, f"book{i}")
            self.assertEqual(sheet.cell(row=(i+1), column=2).value, f"genre{i}")
            self.assertEqual(sheet.cell(row=(i+1), column=3).value, f"${i}.00")

    def test_excel_column_headers(self):
        """Check if column headers match expectations using pandas."""
        df = pd.read_excel(FILENAME)
        self.assertEqual(list(df.columns), ["Title", "Genre", "Price"])

    @classmethod
    def tearDownClass(self):
        """Runs once after all tests finish."""
        if os.path.exists(FILENAME):
            os.remove(FILENAME)

if __name__ == '__main__':
    unittest.main()